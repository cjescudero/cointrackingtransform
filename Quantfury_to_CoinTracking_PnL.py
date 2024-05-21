import pandas as pd
import warnings
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import os

# Suprimir advertencias de openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

# Define the directory containing the files
directory = 'History from Quantfury'

# Initialize an empty DataFrame to store all the data
all_data = pd.DataFrame()

# Iterate over all files in the directory
for filename in os.listdir(directory):
    if filename.endswith(".xlsx"):
        print(filename)
        file_path = os.path.join(directory, filename)
        data = pd.read_excel(file_path)
        all_data = pd.concat([all_data, data], ignore_index=True)

# Filter out rows where "Total Position PnL" is not NaN
filtered_data = all_data.dropna(subset=["Total Position PnL"]).copy()

# Function to clean and convert "Total Position PnL" to float
def clean_pnl(pnl):
    if isinstance(pnl, str):
        pnl = pnl.replace("$", "").replace("₮", "")
    try:
        return float(pnl)
    except ValueError:
        return None

# Clean the "Total Position PnL" column using .loc to avoid SettingWithCopyWarning
filtered_data.loc[:, "Total Position PnL"] = filtered_data["Total Position PnL"].apply(clean_pnl)

# Drop rows where "Total Position PnL" could not be converted to float
filtered_data = filtered_data.dropna(subset=["Total Position PnL"])

# Updated transformation function
def transform_row(row):
    pnl = row["Total Position PnL"]  # Already cleaned and converted to float
    date = pd.to_datetime(row["Date"], format="%d.%m.%Y %I:%M %p UTC").strftime("%d-%m-%Y %H:%M:%S")
    trade_group = "Long" if row["Action"] == "Sold" else "Short"
    
    #print(f'{date} -> {pnl}')

    base_data = {
        "Type": "",
        "Buy Amount": None,
        "Buy Cur.": None,
        "Sell Amount": None,
        "Sell Cur.": None,
        "Fee Amount": None,
        "Fee Cur.": None,
        "Exchange (optional)": "Quantfury",
        "Trade Group (optional)": trade_group,
        "Comment (optional)": row["Name"],
        "Date": date,
        "Liquidity pool (optional)": None,
        "Tx-ID (optional)": None,
        "Buy Value in Account Currency (optional)": None,
        "Sell Value in Account Currency (optional)": None
    }
    
    if pnl > 0:
        base_data.update({
            "Type": "Derivatives / Futures Profit",
            "Buy Amount": pnl,
            "Buy Cur.": "USDT"
        })
    else:
        base_data.update({
            "Type": "Derivatives / Futures Loss",
            "Sell Amount": abs(pnl),
            "Sell Cur.": "USDT"
        })
    
    return base_data

# Apply the transformation to the cleaned data
transformed_data = filtered_data.apply(transform_row, axis=1)

# Convert the transformed data to a DataFrame
result_df = pd.DataFrame(transformed_data.tolist())

# Sort the DataFrame by date (ascending order)
result_df["Date"] = pd.to_datetime(result_df["Date"], format="%d-%m-%Y %H:%M:%S")
result_df = result_df.sort_values(by="Date")

# Create a new workbook and add a worksheet
wb = Workbook()
ws = wb.active

# Add the merged header row
header_text = "CoinTracking · Trade Table"
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(result_df.columns))
ws.cell(row=1, column=1).value = header_text

# Append the DataFrame to the worksheet
for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), start=2):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Save the workbook to a file
output_file_path = 'Quantfury_CoinTracking_data_PnL.xlsx'
wb.save(output_file_path)

print(f'Data saved to {output_file_path}')
